import datetime
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter

from utils.get_stock_history_data import stock_data_gain

redFill = PatternFill(start_color='EE6363',
                      end_color='EE6363',
                      fill_type='solid')
yellowFill = PatternFill(start_color='FFCC11',
                      end_color='FFCC11',
                      fill_type='solid')

blueFill = PatternFill(start_color='00B2EE',
                      end_color='00B2EE',
                      fill_type='solid')

greenFill = PatternFill(start_color='5B965B',
                      end_color='5B965B',
                      fill_type='solid')

'''
    input: cell
    return: column_letter, row_index
'''
def get_cell_coordinate(cell):
    xy = coordinate_from_string(cell.coordinate)

    return xy[0], xy[1]

'''
    input:
        ws: worksheet
        row: row index. default is ':', meaning the whole row
        col: col letter. default is ':', meaning the whole col
        colorFill
    raise error: ValueError
    return: None
'''
def highlight(ws, row, col, colorFill):
    index = ''
    if row == ':' and col == ':':
        raise ValueError
    elif row == ':':
        index += col
    elif col == ':':
        index += row
    else:
        index = col + row

    index = index + ':' + index

    for cell in ws[index]:
        cell.fill = colorFill

def hightlight_row(ws, colName, colorFill, fun):
    if colName is None:
        raise ValueError

    # find col letter with colName
    col_letter = ""
    for cell in ws['1:1']:
        if cell.value == colName:
            col_letter, _ = get_cell_coordinate(cell)

    index = col_letter + ':' + col_letter
    # loop over the column
    for cell in ws[index]:
        if fun(cell.value):
            #get the row
            col, row = get_cell_coordinate(cell)
            # color the row
            highlight(ws, str(row), ':', colorFill)

def hightlight_signalchange(book, sheetTitle, colName):
    sheet = None
    prev1Sheet = None
    prev2Sheet = None

    for ws in book.worksheets:
        prev2Sheet = prev1Sheet
        prev1Sheet = sheet
        sheet = ws

        if ws.title == sheetTitle:
            sheet = ws
            break

    # find col letter with colName
    col_letter = ""
    for cell in sheet['1:1']:
        if cell.value == colName:
            col_letter, _ = get_cell_coordinate(cell)

    index = col_letter + ':' + col_letter
    # loop over the column
    for cell in sheet[index]:
        value1 = cell.value
        col, row = get_cell_coordinate(cell)
        value0 = prev2Sheet[col+str(row)].value

        if value1 == 'Up' and value0 == 'Flat':
            cell.fill = yellowFill

        if value1 == 'Flat' and value0 == 'Up':
            cell.fill = blueFill

def add_new_columns(ws, colNames, last_col_name):
    # add startPrice endPrice and Gain
    row = ws['1:1']

    last_col_index = 0
    for cell in row:
        if cell.value == last_col_name:
            col_letter, _ = get_cell_coordinate(cell)
            last_col_index = column_index_from_string(col_letter)

    if last_col_index == 0:
        raise ValueError

    new_column_letters = []

    for i, name in enumerate(colNames):
        new_col_letter = get_column_letter(last_col_index + i + 1)
        new_column_letters += new_col_letter
        ws['%s%s' % (new_col_letter, '1')].value = name

    return new_column_letters

def fill_initial_process(ws, colNames, last_col_name):
    new_letters = add_new_columns(ws, colNames, last_col_name)
    symbol_index = 2
    p_index = 3
    call_index = 5
    put_index = 6
    comb_index = 7
    index = 2
    for row in ws.iter_rows(min_row=2):
        symbol = row[symbol_index].value
        print ('Processing symbol: %s initial_processing' % symbol)
        gain1 = (row[comb_index].value - row[p_index].value) / row[p_index].value
        gain2 = (row[call_index].value - row[p_index].value) / row[p_index].value
        up_rate = np.cbrt(np.abs((row[p_index].value - row[put_index].value) / (row[call_index].value - row[put_index].value)) - 0.5) * 0.5 + 0.5
        ws['%s%s' % (new_letters[0], str(index))] = gain1
        ws['%s%s' % (new_letters[1], str(index))] = gain2
        ws['%s%s' % (new_letters[2], str(index))] = up_rate
        index += 1

def fill_in_gain(ws, colNames, last_col_name, startDate, endDate):
    new_letters = add_new_columns(ws, colNames, last_col_name)
    symbol_index = 2

    data = []
    # loop row by row
    index = 2
    for row in ws.iter_rows(min_row=2):
        symbol = row[symbol_index].value
        print ('Processing symbol: %s' % symbol)
        data = stock_data_gain(symbol, startDate, endDate)
        ws['%s%s' % (new_letters[0], str(index))] = data[0]
        ws['%s%s' % (new_letters[1], str(index))] = data[1]
        ws['%s%s' % (new_letters[2], str(index))] = data[2]
        index += 1

def understand_excel(Title = None):
    pwd = os.getcwd()
    filename = pwd + '/data/EstimateResult.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    if os.path.isfile(filename):
        # load existing excel
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    if Title is None:
        #find the sheet with title sheetTitle
        sheetTitle = str(datetime.date.today())
    else:
        sheetTitle = Title

    sheet = writer.book.get_sheet_by_name(sheetTitle)
    if sheet is None:
        print("Cannot find the right sheet")

    fill_initial_process(sheet, ['gain1', 'gain2', 'up-rate'], 'Trend_1')

    startDate = str(datetime.date(2017, 11, 27))
    endDate = str(datetime.date.today())
    fill_in_gain(sheet, ['StartPrice', 'EndPrice', 'Gain'], 'up-rate', startDate, endDate)

    # highlight Up
    hightlight_row(sheet, 'Trend_0', greenFill, lambda x: True if x == 'Up' else False)
    # hightlight signal change
    hightlight_signalchange(book, sheetTitle, 'Trend_0')

    writer.save()

if __name__ == '__main__':
    #load excel
    pwd = os.getcwd()
    filename = pwd + '/data/EstimateResult.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    if os.path.isfile(filename):
        # load existing excel
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    sheetTitle = str(datetime.date.today())

    #find the sheet with title sheetTitle
    sheet = writer.book.get_sheet_by_name(sheetTitle)
    if sheet is None:
        print("Cannot find the right sheet")

    writer.save()

    understand_excel(sheetTitle)
