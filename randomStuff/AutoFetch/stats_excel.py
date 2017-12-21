import datetime
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
import matplotlib.pyplot as plt
import sys, getopt

from utils.get_stock_history_data import stock_data_gain, get_stock_price_from_excel, get_stock_history_data2

def parseArg(argv):
   startIndex = 0
   try:
      opts, args = getopt.getopt(argv,"hs:")
   except getopt.GetoptError:
      print ('test.py -s <startIndex>')
      sys.exit(2)
   for opt, arg in opts:
      if opt == '-h':
         print ('test.py -s <startIndex>')
         sys.exit()
      elif opt == '-s':
         startIndex = int(arg)

   print ('start index is %s' % startIndex)

   return startIndex

def get_cell_coordinate(cell):
    xy = coordinate_from_string(cell.coordinate)

    return xy[0], xy[1]

def get_column_letter_by_colname(ws, colname):
    row = ws['1:1']

    col_letter = ''
    col_index = 0
    for cell in row:
        if cell.value == colname:
            col_letter, _ = get_cell_coordinate(cell)
            col_index = column_index_from_string(col_letter)

    if col_letter == '':
        raise ValueError

    return col_letter, col_index

def get_data_from_excel(ws, colNames):
    col_letters = []
    col_index = []
    symbols = []

    for name in colNames:
        letter, index = get_column_letter_by_colname(ws, name)
        col_letters += letter
        col_index.append(index)

    # get data row by row
    data = []
    for row in ws.iter_rows(min_row=2, min_col = col_index[0], max_col = col_index[-1]):
        aRow = []
        for cell in row:
            aRow.append(cell.value)
        data.append(aRow)

    data = np.array(data)

    for cell in ws['c:c']:
        if cell.value == 'Symbol':
            continue
        symbols.append(cell.value)

    return symbols, data

if __name__ == '__main__':
    #load excel
    pwd = os.getcwd()
    filename = pwd + '/data/EstimateResult.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    if os.path.isfile(filename):
        # load existing excel
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    startDate = datetime.date(2017, 11, 20)
    targetDate = datetime.date(2017, 12, 15)
    sheetTitle = str(startDate)

    #find the sheet with title sheetTitle
    sheet = writer.book.get_sheet_by_name(sheetTitle)
    if sheet is None:
        print("Cannot find the right sheet")

    colNames = ['callEst_0', 'putEst_0', 'combEst_0']
    symbols, data = get_data_from_excel(sheet, colNames)
    prices = np.zeros((len(symbols), 3))

    pwd = os.getcwd()
    foldername = pwd + '/data/history/'

    startIndex = parseArg(sys.argv[1:])

    for i, symbol in enumerate(symbols):
        start = startDate
        end = targetDate
        excel_name = foldername + symbol + '_history.xlsx'
        info = get_stock_history_data2(symbol, start, end)
        p = 0
        if info is not None:
            p_array = info['Close'].as_matrix()
            prices[i, 0] = p_array[-1]
            prices[i, 1] = p_array[0]
            prices[i, 2] = (p_array[startIndex] - p_array[-1]) / p_array[-1]

    symbols = np.array(symbols)
    data = np.hstack((symbols.reshape((-1, 1)), data, prices))

    judge = []
    for i in range(len(symbols)):
        callEst = float(data[i, 1].replace(',', ''))
        putEst = float(data[i, 2].replace(',', ''))
        endPrice = float(data[i, 5].replace(',', ''))
        startPrice = float(data[i, 4].replace(',', ''))
        tmp = 'up' if endPrice > startPrice * 1.03 else 'dw'
        '''
        if endPrice > callEst * 0.99:
            tmp = 'up'
        elif endPrice < putEst * 1.01:
            tmp = 'dw'
        '''
        judge.append(tmp)

    judge = np.array(judge)
    data = np.hstack((data, judge.reshape((-1, 1))))

    midNum = np.sum(data[:, 7] == 'mid')

    names = ['symbol', 'callEst', 'putEst', 'combEst', '11-20Price', '12-15Price', 'First5DayCP', 'Direction']
    df = pd.DataFrame(data, columns=names)
    df.to_csv('data/option-11-20To12-15.csv', index=True, header=True, sep=' ')

    #print (data)
    #print (1 - midNum / len(symbols))

'''
    
    price_of_day = np.array(price_of_day)

    Y = np.divide(data[:, 2].reshape((-1, 1)), prices) - 1
    print (Y.shape)

    unit = (data[:, 0] - data[:, 1]) / 2
    zero = (data[:, 0] + data[:, 1]) / 2

    X = np.divide(price_of_day - zero.reshape((-1, 1)), unit.reshape((-1, 1)))
    print (X.shape)

    print (np.column_stack((X, Y)))
    print (prices)
    print (data[:, 2])
    gain = ((data[:, 2].reshape((-1, 1)) - prices) / prices)
    print (symbols)
    print (gain.reshape((1, -1)))
'''

'''
    plt.plot(X, Y, 'ro')
    plt.axis([-2, 2, -2, 2])
    plt.show()
''' 

